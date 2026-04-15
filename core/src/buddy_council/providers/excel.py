"""Excel requirements provider (Jama export fallback)."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from buddy_council.models import Requirement
from buddy_council.normalization import clean_text, normalize_feature_name, normalize_id


# Default column mapping for Jama exports
DEFAULT_UNNAMED_COLUMNS = [
    "ID",
    "Description",
    "Rationale",
    "Item Type",
    "Status",
    "Jira ID",
    "Tags",
    "Configuration",
]


class ExcelProvider:
    """Fetch requirements from a Jama-exported Excel file."""

    def __init__(self, excel_path: str | Path) -> None:
        self._path = Path(excel_path)
        if not self._path.exists():
            raise FileNotFoundError(f"Excel file not found: {self._path}")

    async def fetch(self, scope: str | None = None) -> list[Requirement]:
        """Parse the Excel file and return requirements in canonical schema."""
        # openpyxl is sync — run in the current thread (fine for file I/O)
        from openpyxl import load_workbook

        wb = load_workbook(self._path, read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            return []

        rows = list(ws.iter_rows(min_row=4, values_only=True))  # skip first 3 metadata rows
        if not rows:
            return []

        # Get headers from row 4 (index 0 after skip)
        headers = list(ws.iter_rows(min_row=4, max_row=4, values_only=True))[0]
        data_rows = list(ws.iter_rows(min_row=5, values_only=True))

        # Rename unnamed columns
        column_names = []
        unnamed_idx = 0
        for h in headers:
            h_str = str(h) if h else ""
            if not h_str or "Unnamed" in h_str or h_str == "None":
                if unnamed_idx < len(DEFAULT_UNNAMED_COLUMNS):
                    column_names.append(DEFAULT_UNNAMED_COLUMNS[unnamed_idx])
                    unnamed_idx += 1
                else:
                    column_names.append(h_str)
            else:
                column_names.append(h_str)

        # Build a column index lookup
        col_idx: dict[str, int] = {name: i for i, name in enumerate(column_names)}

        def get_cell(row: tuple, col_name: str) -> str:
            idx = col_idx.get(col_name)
            if idx is None or idx >= len(row):
                return ""
            val = row[idx]
            if val is None:
                return ""
            return str(val).strip()

        # Parse rows, tracking current feature (folder)
        requirements: list[Requirement] = []
        current_feature = "Unknown"

        for row in data_rows:
            item_type = get_cell(row, "Item Type")
            if item_type == "Folder":
                # Use the Name column or first column as feature name
                name_idx = col_idx.get("Name", 0)
                if name_idx < len(row) and row[name_idx]:
                    current_feature = str(row[name_idx]).strip()
                continue

            # Skip non-requirement rows
            if item_type in ("Text", "Folder", ""):
                continue

            req_id = get_cell(row, "ID")
            if not req_id or req_id in ("nan", "None"):
                continue

            requirements.append(
                Requirement(
                    id=normalize_id(req_id),
                    title=get_cell(row, "Name") or get_cell(row, column_names[0]) if column_names else "",
                    description=clean_text(get_cell(row, "Description")),
                    feature=normalize_feature_name(current_feature),
                    status=get_cell(row, "Status"),
                    linked_ids=frozenset(),
                    raw_fields={
                        name: get_cell(row, name)
                        for name in column_names
                        if get_cell(row, name) and get_cell(row, name) not in ("nan", "None")
                    },
                )
            )

        wb.close()

        # Filter by scope
        if scope and scope.lower() != "all":
            normalized_scope = scope.upper()
            if normalized_scope.startswith("CWA-"):
                # Specific requirement ID
                requirements = [r for r in requirements if r.id == normalized_scope]
            else:
                # Feature name
                requirements = [
                    r for r in requirements if scope.lower() in r.feature.lower()
                ]

        return requirements
